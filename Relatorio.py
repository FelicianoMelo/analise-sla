import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, PatternFill

class ExcelReport:

    def __init__(
        self,
        arquivo,
        sheet_name=None,
        linhas_em_branco=2
    ):
        self.arquivo = arquivo
        self.sheet_name = sheet_name
        self.linhas_em_branco = linhas_em_branco

        self.borda = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

    def aplicar_borda(
        self,
        ws,
        linha_inicio,
        n_linhas,
        n_colunas
    ):
        linha_fim = linha_inicio + n_linhas

        for row in ws.iter_rows(
            min_row=linha_inicio,
            max_row=linha_fim,
            min_col=1,
            max_col=n_colunas
        ):
            for cell in row:
                cell.border = self.borda

    def gerar(self, dataframes):
        """
        dataframes: lista de DataFrames
        """

        posicao_linha = 0
        blocos = []

        # Escreve os DataFrames
        with pd.ExcelWriter(
            self.arquivo,
            engine="openpyxl",
            mode="a",
            if_sheet_exists='overlay'
        ) as writer:

            for df in dataframes:

                df.to_excel(
                    writer,
                    sheet_name=self.sheet_name,
                    startrow=posicao_linha,
                    index=False

                )

                blocos.append({
                    "linha_inicio": posicao_linha + 1,
                    "n_linhas": len(df),
                    "n_colunas": len(df.columns)
                })

                posicao_linha += (
                    len(df)
                    + 1  # cabeçalho
                    + self.linhas_em_branco
                )

        # Aplica bordas
        wb = load_workbook(self.arquivo)
        ws = wb[self.sheet_name]

        for bloco in blocos:
            self.aplicar_borda(
                ws,
                bloco["linha_inicio"],
                bloco["n_linhas"],
                bloco["n_colunas"]
            )

        # Ajustar largura das colunas
        for coluna in ws.columns:
            largura = max(
                len(str(cell.value))
                if cell.value is not None else 0
                for cell in coluna
            )

            ws.column_dimensions[
                coluna[0].column_letter
            ].width = largura + 2

        # Definir cores e estilos
        fill_titulo = PatternFill(
            fill_type="solid", start_color="1F4E79", end_color="1F4E79"
        )  # Azul escuro
        font_titulo = Font(color="FFFFFF", bold=True)

        fill_total = PatternFill(
            fill_type="solid", start_color="D9E1F2", end_color="D9E1F2"
        )  # Azul claro
        font_total = Font(bold=True)

        # Bordas com Side e Border
        borda_fina = Side(border_style="thin", color="000000")
        borda_dupla = Side(border_style="double", color="000000")

        borda_titulo = Border(
            top=borda_fina, bottom=borda_fina, left=borda_fina, right=borda_fina
        )
        borda_total = Border(
            top=borda_fina, bottom=borda_dupla, left=borda_fina, right=borda_fina
        )

        # 1. Aplicar estilos na linha do Título (Linha 1)
        for cell in ws[1]:
            cell.fill = fill_titulo
            cell.font = font_titulo
            cell.border = borda_titulo

        # 2. Aplicar estilos na linha do Total (Linha 4)
        total_row = 4
        for cell in ws[total_row]:
            cell.fill = fill_total
            cell.font = font_total
            cell.border = borda_total


        wb.save(self.arquivo)