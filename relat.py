import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter

class ExcelReport:

    def __init__(
        self,
        arquivo,
        sheet_name="Relatório",
        linhas_em_branco=2
    ):
        self.arquivo = arquivo
        self.sheet_name = sheet_name
        self.linhas_em_branco = linhas_em_branco

        # --- Definição de Cores, Fontes e Bordas (Fornecidos no seu script) ---
        self.fill_titulo = PatternFill(
            fill_type="solid", start_color="1F4E79", end_color="1F4E79"
        )  # Azul escuro
        self.font_titulo = Font(color="FFFFFF", bold=True)

        self.fill_total = PatternFill(
            fill_type="solid", start_color="D9E1F2", end_color="D9E1F2"
        )  # Azul claro
        self.font_total = Font(bold=True)

        # Bordas com Side e Border
        borda_fina = Side(border_style="thin", color="000000")
        borda_dupla = Side(border_style="double", color="000000")

        self.borda_titulo = Border(
            top=borda_fina, bottom=borda_fina, left=borda_fina, right=borda_fina
        )
        self.borda_total = Border(
            top=borda_fina, bottom=borda_dupla, left=borda_fina, right=borda_fina
        )
        self.borda_comum = Border(
            top=borda_fina, bottom=borda_fina, left=borda_fina, right=borda_fina
        )

    def _garantir_arquivo_existe(self):
        """Cria o arquivo caso ele não exista para suportar o modo 'a' no pandas."""
        if not os.path.exists(self.arquivo):
            wb = Workbook()
            ws = wb.active
            ws.title = self.sheet_name
            wb.save(self.arquivo)

    def aplicar_estilos_bloco(
        self,
        ws,
        linha_inicio,
        n_linhas,
        n_colunas
    ):
        """
        Aplica os estilos de Título (cabeçalho), Dados e Total (última linha)
        para um bloco específico de dados.
        """
        # A linha final do bloco é: inicio + 1 (cabeçalho) + quantidade de linhas de dados - 1
        linha_fim = linha_inicio + n_linhas

        for row_idx, row in enumerate(ws.iter_rows(
            min_row=linha_inicio,
            max_row=linha_fim,
            min_col=1,
            max_col=n_colunas
        ), start=linha_inicio):
            
            # 1. Caso seja a linha do Título/Cabeçalho (Primeira linha do bloco)
            if row_idx == linha_inicio:
                for cell in row:
                    cell.fill = self.fill_titulo
                    cell.font = self.font_titulo
                    cell.border = self.borda_titulo

            # 2. Caso seja a linha de Total (Última linha do bloco)
            elif row_idx == linha_fim:
                for cell in row:
                    cell.fill = self.fill_total
                    cell.font = self.font_total
                    cell.border = self.borda_total

            # 3. Linhas de dados comuns do meio
            else:
                for cell in row:
                    cell.border = self.borda_comum

    def gerar(self, dataframes):
        """
        dataframes: lista de DataFrames a serem gravados
        """
        self._garantir_arquivo_existe()

        posicao_linha = 0
        blocos = []

        # 1. Escrever DataFrames no Excel usando Pandas
        with pd.ExcelWriter(
            self.arquivo,
            engine="openpyxl",
            mode="a",
            if_sheet_exists='overlay'
        ) as writer:

            for df in dataframes:
                df.to_excel(
                    writer,
                    sheet_name=self.sheet_name,
                    startrow=posicao_linha,
                    index=False
                )

                blocos.append({
                    "linha_inicio": posicao_linha + 1,  # OpenPyXL usa índice base 1
                    "n_linhas": len(df),
                    "n_colunas": len(df.columns)
                })

                posicao_linha += (
                    len(df)
                    + 1  # Cabeçalho
                    + self.linhas_em_branco
                )

        # 2. Aplicar Formatações com OpenPyXL
        wb = load_workbook(self.arquivo)
        ws = wb[self.sheet_name]

        # Aplica estilos para cada tabela escrita
        for bloco in blocos:
            self.aplicar_estilos_bloco(
                ws,
                bloco["linha_inicio"],
                bloco["n_linhas"],
                bloco["n_colunas"]
            )

        # 3. Auto-ajuste da largura das colunas
        for col_idx, coluna in enumerate(ws.columns, start=1):
            largura_maxima = 0
            for cell in coluna:
                if cell.value is not None:
                    largura_maxima = max(largura_maxima, len(str(cell.value)))

            letra_coluna = get_column_letter(col_idx)
            ws.column_dimensions[letra_coluna].width = max(largura_maxima + 4, 12)

        wb.save(self.arquivo)
