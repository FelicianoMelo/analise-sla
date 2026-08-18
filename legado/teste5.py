from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, PatternFill
def aplicar_estilos_excel(arquivo_saida,aba):
    # 1. Carregar o arquivo Excel existente
    wb = load_workbook(arquivo_saida)
    ws = wb[aba]

    # Definir cores e estilos
    fill_titulo = PatternFill(
        fill_type="solid", start_color="1F4E79", end_color="1F4E79"
    )  # Azul escuro
    font_titulo = Font(color="FFFFFF", bold=True)

    fill_total = PatternFill(
        fill_type="solid", start_color="D9E1F2", end_color="D9E1F2"
    )  # Azul claro
    font_total = Font(bold=True)

    # Bordas com Side e Border
    borda_fina = Side(border_style="thin", color="000000")
    borda_dupla = Side(border_style="double", color="000000")

    borda_titulo = Border(
        top=borda_fina, bottom=borda_fina, left=borda_fina, right=borda_fina
    )


    # Aplicar estilos na linha do Título (Linha 1)
    for cell in ws[1]:
                cell.fill = fill_titulo
                cell.font = font_titulo
                cell.border = borda_titulo

    # Salvar o arquivo
    wb.save(arquivo_saida)
