import pandas as pd
import json
import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Data atual para salvar o arquivo com um nome único
data_atual = datetime.datetime.now().strftime('%Y%m%d')

def cruzar_planilhas(arquivo1, arquivo2, arquivo_saida):
    # 1. Ler os dados das planilhas para DataFrames
    df_plan1 = pd.read_excel(arquivo1, data_only=True)
    df_plan2 = pd.read_excel(arquivo2, data_only=True)

    # 2. Renomear a coluna 0 para 'Chave' (caso o nome seja diferente)
    df_plan1.rename(columns={df_plan1.columns[0]: 'Chave'}, inplace=True)
    df_plan2.rename(columns={df_plan2.columns[0]: 'Chave'}, inplace=True)

    # 3. Mesclar (Unir) os dados com base na coluna 'Chave'
    df_plan3 = pd.merge(df_plan1, df_plan2, on='Chave', how='left')

    # 4. Salvar o resultado e aplicar formatação com openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado Cruzamento"

    # Converter o DataFrame para linhas do openpyxl
    for r_idx in dataframe_to_rows(df_plan3, index=False, header=True):
        ws.append(r_idx)

    # --- CONFIGURAÇÃO DE ESTILOS ---
    # Estilo do Cabeçalho: Azul escuro, texto branco e negrito
    cor_cabecalho = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fonte_cabecalho = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    # Estilo dos Dados: Fonte padrão
    fonte_dados = Font(name="Calibri", size=11, bold=False)
    
    # Alinhamento centralizado geral
    alinhamento = Alignment(horizontal="center", vertical="center")
    
    # Bordas finas cinzas para todas as células
    borda_fina = Side(border_style="thin", color="D9D9D9")
    estilo_borda = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)

    # --- APLICAR ESTILOS NAS CÉLULAS ---
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=1):
        for cell in row:
            cell.alignment = alinhamento
            cell.border = estilo_borda
            
            if row_idx == 1:
                # Aplicar estilo específico de cabeçalho na primeira linha
                cell.fill = cor_cabecalho
                cell.font = fonte_cabecalho
            else:
                # Aplicar estilo de dados nas demais linhas
                cell.font = fonte_dados

    # Ajustar a largura das colunas automaticamente baseado no conteúdo
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Salvar o arquivo final formatado
    wb.save(arquivo_saida)

    # 5. Exibir mensagem de sucesso
    print(f"Arquivo {arquivo_saida} criado e formatado com sucesso!")

def processar_configuracoes():
    # Ler o arquivo JSON
    with open('configuracao.json', 'r', encoding='utf-8') as f:
        configuracoes = json.load(f)

    # Criar uma nova planilha com base nas configurações do arquivo JSON
    for config in configuracoes:
        arquivo1 = config['ArquivoEntrada1']
        arquivo2 = config['ArquivoEntrada2']
        arquivo_saida = f"{data_atual}_{config['arquivoSaida']}"
        
        # Chame a função para cruzar as planilhas
        cruzar_planilhas(arquivo1, arquivo2, arquivo_saida)


def main():
    processar_configuracoes()     

if __name__ == "__main__":
    main()