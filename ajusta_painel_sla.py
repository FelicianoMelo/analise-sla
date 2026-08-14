from openpyxl.styles import (
    PatternFill,
    Font,
    Border,
    Side,
    Alignment
)

BORDA = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)


def escrever_celula(
        ws,
        linha,
        col_ini,
        col_fim,
        texto,
        cor=None,
        fonte_negrito=True,
        alinhamento="center"
):

    ws.merge_cells(
        start_row=linha,
        start_column=col_ini,
        end_row=linha,
        end_column=col_fim
    )

    cel = ws.cell(linha, col_ini)
    cel.value = texto
    cel.font = Font(bold=fonte_negrito)

    if cor:
        cel.fill = PatternFill(
            fill_type="solid",
            fgColor=cor
        )

    cel.alignment = Alignment(
        horizontal=alinhamento,
        vertical="center"
    )

    for c in range(col_ini, col_fim + 1):
        ws.cell(linha, c).border = BORDA


def criar_painel_sla(
        ws,
        coluna_inicial,
        titulo,
        data_aberto,
        data_encerrado,
        atribuicoes,
        resolvidos,
        sla_otimo,
        sla_medio,
        sla_atencao,
        sla_fora,
        detalhe_otimo=None,
        detalhe_medio=None,
        detalhe_atencao=None,
        detalhe_fora=None
):

    c = coluna_inicial

    detalhe_otimo = detalhe_otimo or {}
    detalhe_medio = detalhe_medio or {}
    detalhe_atencao = detalhe_atencao or {}
    detalhe_fora = detalhe_fora or {}

    # ==================================================
    # TÍTULO
    # ==================================================

    escrever_celula(
        ws,
        2,
        c,
        c + 7,
        titulo,
        "C4D96D"
    )

    # ==================================================
    # CABEÇALHO
    # ==================================================

    escrever_celula(
        ws,
        4,
        c,
        c + 3,
        f"em Aberto: {data_aberto}",
        "D9D9D9"
    )

    escrever_celula(
        ws,
        4,
        c + 4,
        c + 7,
        f"Encerrados: {data_encerrado}",
        "D9D9D9"
    )

    # ==================================================
    # ABERTOS
    # ==================================================

    linha_aberto = 6

    ws.cell(linha_aberto, c, "Atribuído a")
    ws.cell(linha_aberto, c + 3, "Quantidade")

    for col in [c, c + 1, c + 2, c + 3]:
        ws.cell(linha_aberto, col).fill = PatternFill(
            "solid",
            fgColor="F2C94C"
        )
        ws.cell(linha_aberto, col).font = Font(bold=True)

    linha_aberto += 1

    total_abertos = 0

    for nome, qtd in atribuicoes.items():

        ws.cell(linha_aberto, c, nome)
        ws.cell(linha_aberto, c + 3, qtd)

        total_abertos += qtd
        linha_aberto += 1

    ws.cell(linha_aberto, c, "TOTAL")
    ws.cell(linha_aberto, c + 3, total_abertos)

    # ==================================================
    # ENCERRADOS
    # ==================================================

    linha_enc = 6

    ws.cell(linha_enc, c + 5, "Resolvido por")
    ws.cell(linha_enc, c + 7, "Quantidade")

    for col in [c + 5, c + 6, c + 7]:
        ws.cell(linha_enc, col).fill = PatternFill(
            "solid",
            fgColor="E6E6A8"
        )
        ws.cell(linha_enc, col).font = Font(bold=True)

    linha_enc += 1

    total_enc = 0

    for nome, qtd in resolvidos.items():

        ws.cell(linha_enc, c + 5, nome)
        ws.cell(linha_enc, c + 7, qtd)

        total_enc += qtd
        linha_enc += 1

    ws.cell(linha_enc, c + 5, "TOTAL")
    ws.cell(linha_enc, c + 7, total_enc)

    # ==================================================
    # SLA TOTAL
    # ==================================================

    linha = max(linha_aberto, linha_enc) + 5

    linha_sla_total = linha

    escrever_celula(
        ws,
        linha_sla_total,
        c,
        c + 3,
        f"SLA - TOTAL {total_abertos}",
        "B7D7F0"
    )

    # ==================================================
    # Dentro do SLA
    # ==================================================

    linha += 2

    linha_sla_otimo = linha

    escrever_celula(
        ws,
        linha_sla_otimo,
        c,
        c + 3,
        f"Dentro do SLA: {sla_otimo}",
        "98FB98"
    )

    linha_detalhe = linha_sla_otimo + 1

    for fila, qtd in detalhe_otimo.items():

        ws.cell(linha_detalhe, c + 4, fila)
        ws.cell(linha_detalhe, c + 7, qtd)

        linha_detalhe += 1

    # ==================================================
    # SLA MÉDIO
    # ==================================================

    linha = linha_detalhe + 2

    linha_sla_medio = linha

    escrever_celula(
        ws,
        linha_sla_medio,
        c,
        c + 3,
        f"SLA Médio: {sla_medio}",
        "D4AF37"
    )

    linha_detalhe = linha_sla_medio + 1

    for fila, qtd in detalhe_medio.items():

        ws.cell(linha_detalhe, c + 4, fila)
        ws.cell(linha_detalhe, c + 7, qtd)

        linha_detalhe += 1

    # ==================================================
    # SLA EM ATENÇÃO
    # ==================================================

    linha = linha_detalhe + 2

    linha_sla_atencao = linha

    escrever_celula(
        ws,
        linha_sla_atencao,
        c,
        c + 3,
        f"SLA em Atenção: {sla_atencao}",
        "F4A460"
    )

    linha_detalhe = linha_sla_atencao + 1

    for fila, qtd in detalhe_atencao.items():

        ws.cell(linha_detalhe, c + 4, fila)
        ws.cell(linha_detalhe, c + 7, qtd)

        linha_detalhe += 1

    # ==================================================
    # FORA DE SLA
    # ==================================================

    linha = linha_detalhe + 2

    linha_sla_fora = linha

    escrever_celula(
        ws,
        linha_sla_fora,
        c,
        c + 3,
        f"Fora de SLA: {sla_fora}",
        "FF3030"
    )

    linha_detalhe = linha_sla_fora + 1

    for fila, qtd in detalhe_fora.items():

        ws.cell(linha_detalhe, c + 4, fila)
        ws.cell(linha_detalhe, c + 7, qtd)

        linha_detalhe += 1

    # ==================================================
    # BORDAS
    # ==================================================

    for row in ws.iter_rows(
        min_row=2,
        max_row=linha_detalhe,
        min_col=c,
        max_col=c + 7
    ):
        for cell in row:
            cell.border = BORDA

    return linha_detalhe