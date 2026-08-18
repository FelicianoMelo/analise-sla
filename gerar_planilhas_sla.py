from datetime import datetime, timedelta
import json
import logging
from pathlib import Path
from typing import Dict, Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import pandas as pd

from relat import ExcelReport
import ajusta_painel_sla as aps


# ==============================================================================
# CONFIGURAÇÃO DE AMBIENTE E LOGGING
# ==============================================================================
PASTA_TRABALHO = Path.home()  / "projetos/analise-sla/"
print(PASTA_TRABALHO)
logger = logging.getLogger("SLA")
logger.setLevel(logging.INFO)

formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")

# Handler de arquivo
arquivo_log = logging.FileHandler(PASTA_TRABALHO / "analise_sla.log", encoding="utf-8")
arquivo_log.setFormatter(formatter)

# Handler de console
console = logging.StreamHandler()
console.setFormatter(formatter)

if not logger.handlers:
    logger.addHandler(arquivo_log)
    logger.addHandler(console)

# Constantes de data
DATA_ATUAL = datetime.now().strftime("%Y%m%d")
DATA_ONTEM = (datetime.now() - timedelta(days=1)).strftime("%Y%m%d")


# ==============================================================================
# FUNÇÕES DE MANIPULAÇÃO DE DATAFRAMES
# ==============================================================================
def agrupar_por(df: pd.DataFrame, coluna: str) -> pd.DataFrame:
    """Agrupa o DataFrame pela coluna especificada e adiciona a linha 'TOTAL GERAL'."""
    if coluna not in df.columns:
        logger.warning(f"Coluna '{coluna}' não encontrada no DataFrame.")
        return pd.DataFrame(columns=[coluna, "Total"])

    resultado = df.groupby(coluna).size().reset_index(name="Total")
    linha_total = pd.DataFrame({
        coluna: ["TOTAL GERAL"],
        "Total": [resultado["Total"].sum()]
    })

    return pd.concat([resultado, linha_total], axis=0, ignore_index=True)


# ==============================================================================
# FUNÇÕES DE ESTILIZAÇÃO E FORMATO DO EXCEL
# ==============================================================================
def aplicar_estilos_excel(caminho_arquivo: Path, aba: str) -> None:
    """Aplica o cabeçalho estilizado (Azul escuro com texto branco) na primeira linha."""
    wb = load_workbook(caminho_arquivo)
    if aba not in wb.sheetnames:
        wb.close()
        return

    ws = wb[aba]

    fill_titulo = PatternFill(fill_type="solid", start_color="1F4E79", end_color="1F4E79")
    font_titulo = Font(color="FFFFFF", bold=True)
    borda_fina = Side(border_style="thin", color="000000")
    borda_titulo = Border(top=borda_fina, bottom=borda_fina, left=borda_fina, right=borda_fina)

    for cell in ws[1]:
        cell.fill = fill_titulo
        cell.font = font_titulo
        cell.border = borda_titulo

    wb.save(caminho_arquivo)
    wb.close()


def colocar_cores_celulas(caminho_arquivo: Path, nome_aba: str, nome_coluna: str) -> Dict[str, int]:
    """Aplica cores nas células baseando-se na meta de SLA e contabiliza a distribuição."""
    wb = load_workbook(caminho_arquivo)
    ws = wb[nome_aba]

    valores_sla = {"otimo": 0, "medio": 0, "atencao": 0, "fora": 0, "TOTAL": 0}

    cores = {
        "verde": PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE"),
        "amarelo": PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C"),
        "laranja": PatternFill("solid", start_color="FFD699", end_color="FFD699"),
        "vermelho": PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE"),
    }

    col_idx = next((cell.column for cell in ws[1] if cell.value == nome_coluna), None)

    if col_idx is None:
        wb.close()
        raise ValueError(f"Coluna '{nome_coluna}' não encontrada na aba '{nome_aba}'.")

    for row in range(2, ws.max_row + 1):
        celula = ws.cell(row=row, column=col_idx)
        valor = celula.value

        if valor is None:
            continue

        if valor <= 45:
            celula.fill = cores["verde"]
            valores_sla["otimo"] += 1
        elif valor <= 50:
            celula.fill = cores["amarelo"]
            valores_sla["medio"] += 1
        elif valor <= 99:
            celula.fill = cores["laranja"]
            valores_sla["atencao"] += 1
        else:
            celula.fill = cores["vermelho"]
            valores_sla["fora"] += 1

    valores_sla["TOTAL"]= valores_sla["otimo"] + valores_sla["medio"] + valores_sla["atencao"] + valores_sla["fora"]
    wb.save(caminho_arquivo)
    wb.close()

    logger.info(f"Cores aplicadas com sucesso no arquivo: {caminho_arquivo.name}")
    return valores_sla


def aplicar_cores_sla_totais(caminho_arquivo: Path, aba: str, valores_sla: Dict[str, int]) -> None:
    """Preenche a tabela resumo de SLA na aba 'SLA'."""
    wb = load_workbook(caminho_arquivo)

    if "SLA" in wb.sheetnames:
        ws = wb["SLA"]
    else:
        ws = wb.worksheets[1]
        ws.title = "SLA"

    if aba in {"INT_aberto", "DOM_aberto", "CSATH_aberto"}:
        _aplicar_sla_resumo(ws, valores_sla)
    else:
        logger.warning(f"Aba '{aba}' não é elegível para o resumo SLA.")

    wb.save(caminho_arquivo)
    wb.close()
    logger.info(f"Resumo SLA atualizado com sucesso no arquivo '{caminho_arquivo.name}'")


def _aplicar_sla_resumo(ws, valores_sla: Dict[str, int]) -> None:
    """Função auxiliar para estruturar os dados da tabela de resumo de SLA."""
    borda = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    fill_cabecalho = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    font_cabecalho = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    ultima_linha = ws.max_row + 3

    letraA = f"A{ultima_linha}"
    letraB = f"B{ultima_linha}"
    # Cabeçalho
    for col, texto in [(letraA, "SLA"), (letraB, "Total")]:
        ws[col].value = texto
        ws[col].fill = fill_cabecalho
        ws[col].font = font_cabecalho
        ws[col].alignment = align_center


    configuracoes = [
        (ultima_linha + 1, "Dentro do SLA", valores_sla.get("otimo", 0), "C6EFCE"),
        (ultima_linha + 2, "Em andamento", valores_sla.get("medio", 0), "FFEB9C"),
        (ultima_linha + 3, "Em atenção", valores_sla.get("atencao", 0), "FFD699"),
        (ultima_linha + 4, "Fora do SLA", valores_sla.get("fora", 0), "FFC7CE"),
        (ultima_linha + 5, "TOTAL GERAL", valores_sla.get("TOTAL", 0), "FFFFFF"),
    ]

    for linha, texto, valor, cor in configuracoes:
        fill_linha = PatternFill(start_color=cor, end_color=cor, fill_type="solid")

        ws[f"A{linha}"].value = texto
        ws[f"B{linha}"].value = valor

        for col_letter in ["A", "B"]:
            cell = ws[f"{col_letter}{linha}"]
            cell.fill = fill_linha
            cell.alignment = align_center

    for row in ws.iter_rows(min_row=ultima_linha +1 , max_row=ultima_linha + 5 , min_col=1, max_col=2):
        for cell in row:
            cell.border = borda
            
# ==============================================================================
# GERAÇÃO DE RELATÓRIOS
# ==============================================================================
def gerar_relatorio_encerrados(caminho_arquivo: Path, aba1: str, aba2: str, df_filtrado: pd.DataFrame) -> None:
    relatorio = ExcelReport(arquivo=str(caminho_arquivo), sheet_name=aba2, linhas_em_branco=2)
    df_agrupa1 = agrupar_por(df_filtrado, "IC afetado")
    df_agrupa2 = agrupar_por(df_filtrado, "Resolvido por")

    relatorio.gerar([df_agrupa1, df_agrupa2])
    aplicar_estilos_excel(caminho_arquivo, aba1)


def gerar_relatorio_aberto(caminho_arquivo: Path, aba1: str, aba2: str, df_filtrado: pd.DataFrame) -> None:
    relatorio = ExcelReport(arquivo=str(caminho_arquivo), sheet_name=aba2, linhas_em_branco=2)
    df_agrupa1 = agrupar_por(df_filtrado, "IC afetado")
    df_agrupa2 = agrupar_por(df_filtrado, "Atribuição a")

    relatorio.gerar([df_agrupa1, df_agrupa2])
    aplicar_estilos_excel(caminho_arquivo, aba1)


def gerar_relatorio_aberto_sctask(caminho_arquivo: Path, aba1: str, aba2: str, df_filtrado: pd.DataFrame) -> None:
    relatorio = ExcelReport(arquivo=str(caminho_arquivo), sheet_name=aba2, linhas_em_branco=2)
    df_agrupa1 = agrupar_por(df_filtrado, "Grupo de atribuição")
    df_agrupa2 = agrupar_por(df_filtrado, "Atribuição a")

    relatorio.gerar([df_agrupa1, df_agrupa2])
    aplicar_estilos_excel(caminho_arquivo, aba1)


# ==============================================================================
# PROCESSAMENTO DE PLANILHAS
# ==============================================================================
def cruzar_planilhas(arquivo1: Path, arquivo2: Path, arquivo_saida: Path, filtro: str, aba1: str, aba2: str) -> None:
    df_plan1 = pd.read_excel(arquivo1)
    df_plan2 = pd.read_excel(arquivo2)

    df_plan1.rename(columns={df_plan1.columns[0]: 'Numero', df_plan1.columns[2]: 'Aberto por'}, inplace=True)
    df_plan2.rename(columns={df_plan2.columns[0]: 'Numero', df_plan2.columns[6]: '% SLA Resolution'}, inplace=True)

    df_plan2 = df_plan2.iloc[:, [6, 0]]
    df_filtrado = df_plan1[df_plan1.iloc[:, 5].astype(str).str.contains(filtro, na=False, regex=True)]

    df_plan3 = pd.merge(df_plan2, df_filtrado, on='Numero', how='right')
    df_plan3.insert(loc=0, column='% SLA Response', value='')
    df_plan3['% SLA Resolution'] = df_plan3['% SLA Resolution'].fillna(0)

    with pd.ExcelWriter(arquivo_saida, engine='openpyxl') as writer:
        df_plan3.to_excel(writer, sheet_name=aba1, index=False)
        pd.DataFrame().to_excel(writer, sheet_name=aba2, index=False)

    aplicar_estilos_excel(arquivo_saida, aba1)
    logger.info(f"Arquivo '{arquivo_saida.name}' cruzado com sucesso!")

    gerar_relatorio_aberto(arquivo_saida, aba1, aba2, df_plan3)


def planilhas_encerrados(arquivo1: Path, arquivo_saida: Path, filtro: str, aba1: str, aba2: str) -> None:
    df_plan1 = pd.read_excel(arquivo1)
    df_plan1.rename(columns={df_plan1.columns[2]: 'Aberto por'}, inplace=True)

    df_filtrado = df_plan1[df_plan1.iloc[:, 5] == filtro]

    with pd.ExcelWriter(arquivo_saida, engine='openpyxl') as writer:
        df_filtrado.to_excel(writer, sheet_name=aba1, index=False)
        pd.DataFrame().to_excel(writer, sheet_name=aba2, index=False)

    aplicar_estilos_excel(arquivo_saida, aba1)
    logger.info(f"Arquivo '{arquivo_saida.name}' gerado com sucesso!")

    gerar_relatorio_encerrados(arquivo_saida, aba1, aba2, df_filtrado)


def planilhas_encerrados_aberto_sctask(arquivo1: Path, arquivo_saida: Path, aba1: str, aba2: str) -> None:
    df_plan1 = pd.read_excel(arquivo1)

    with pd.ExcelWriter(arquivo_saida, engine='openpyxl') as writer:
        df_plan1.to_excel(writer, sheet_name=aba1, index=False)
        pd.DataFrame().to_excel(writer, sheet_name=aba2, index=False)

    gerar_relatorio_aberto_sctask(arquivo_saida, aba1, aba2, df_plan1)
    aplicar_estilos_excel(arquivo_saida, aba1)
    logger.info(f"Arquivo SCTASK '{arquivo_saida.name}' criado com sucesso!")


# ==============================================================================
# FLUXO PRINCIPAL
# ==============================================================================
def processar_configuracoes() -> None:
    caminho_config = PASTA_TRABALHO / "configuracao.json"

    if not caminho_config.exists():
        logger.error(f"Arquivo de configuração não encontrado: {caminho_config}")
        return

    with open(caminho_config, "r", encoding="utf-8") as f:
        configuracoes = json.load(f)

    for item in configuracoes:
        arquivo1 = PASTA_TRABALHO / item["ArquivoEntrada1"]
        arquivo2 = PASTA_TRABALHO / item["ArquivoEntrada2"]
        arquivo_saida = PASTA_TRABALHO / f"{DATA_ATUAL}_{item['arquivoSaida']}"

        filtro = item["Filtro"]
        junta_arq = item.get("JuntaArq", False)
        aba1 = item["NovaAba1"]
        aba2 = item["NovaAba2"]

        # Estrutura padrão de SLA caso a contagem de células não seja executada
        valores_sla_padrao = {"otimo": 0, "medio": 0, "atencao": 0, "fora": 0, "TOTAL":0}

        if "SCTASK" in arquivo1.name:
            planilhas_encerrados_aberto_sctask(arquivo1, arquivo_saida, aba1, aba2)
            aplicar_cores_sla_totais(arquivo_saida, aba1, valores_sla_padrao)

        elif junta_arq:
            planilhas_encerrados(arquivo1, arquivo_saida, filtro, aba1, aba2)
            # Caso queira aplicar cálculo de cores no modo junta_arq, descomente abaixo:
            # valores_sla = colocar_cores_celulas(arquivo_saida, aba1, '% SLA Resolution')
            aplicar_cores_sla_totais(arquivo_saida, aba1, valores_sla_padrao)

        else:
            cruzar_planilhas(arquivo1, arquivo2, arquivo_saida, filtro, aba1, aba2)
            valores_sla = colocar_cores_celulas(arquivo_saida, aba1, '% SLA Resolution')
            aplicar_cores_sla_totais(arquivo_saida, aba1, valores_sla)


def criar_planilha_total() -> None:
    arquivo_saida = PASTA_TRABALHO / f"{DATA_ATUAL}_Total.xlsx"

    with pd.ExcelWriter(arquivo_saida, engine="openpyxl") as writer:
        pd.DataFrame().to_excel(writer, sheet_name="Visão Geral", index=False)

    wb = load_workbook(arquivo_saida)
    ws = wb["Visão Geral"]

    # Criar painel Transporte Doméstico
    aps.criar_painel_sla(
        ws,
        coluna_inicial=2,
        titulo="TRANSPORTE DOMÉSTICO",
        data_aberto=DATA_ATUAL,
        data_encerrado=DATA_ONTEM,
        atribuicoes={},
        resolvidos={},
        sla_otimo=0,
        sla_medio=0,
        sla_atencao=0,
        sla_fora=0,
    )

    # Criar painel Clearance - Transporte Internacional
    aps.criar_painel_sla(
        ws,
        coluna_inicial=12,
        titulo="CLEARANCE - TRANSPORTE INTERNACIONAL",
        data_aberto=DATA_ATUAL,
        data_encerrado=DATA_ONTEM,
        atribuicoes={
            "Rafael Miranda Ferreira": 18,
            "Enoque Pinheiro": 10,
            "Victorio Bormolini": 1,
            "ANTONIO BARBOSA": 1,
        },
        resolvidos={
            "ANTONIO BARBOSA": 1,
            "Victorio Bormolini": 1,
        },
        sla_otimo=0,
        sla_medio=0,
        sla_atencao=0,
        sla_fora=0,
        detalhe_otimo={"LAC-BR-3-CSATH": 3, "LAC-BR-3-ClearanceFormal": 2},
        detalhe_medio={"LAC-BR-Clearance": 1, "LAC-BR-3-CSATH": 2},
        detalhe_atencao={"LAC-BR-3-CSATH": 5},
        detalhe_fora={"LAC-BR-Clearance": 16, "LAC-BR-3-CSATH": 1},
    )

    wb.save(arquivo_saida)
    wb.close()
    logger.info(f"Planilha resumo final '{arquivo_saida.name}' criada com sucesso!")


def main() -> None:
    logger.info("Iniciando processamento dos relatórios de SLA...")
    processar_configuracoes()
    criar_planilha_total()
    logger.info("Processamento finalizado com sucesso!")


if __name__ == "__main__":
    main()