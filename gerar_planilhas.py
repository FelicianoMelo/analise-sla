import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
import json
import datetime
import os
from relat import ExcelReport 
import ExcelReportHeaderOnly as rel
import ajusta_painel_sla as aps
import logging

data_atual = datetime.datetime.now().strftime("%Y%m%d") # Data atual para salvar o arquivo com um nome único
data_ontem = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime("%Y%m%d") # Data de ontem para salvar o arquivo com um nome único
pasta_trabalho = '/Users/famj/projetos/analise_sla/' # Pasta onde os arquivos serão salvos
os.chdir(pasta_trabalho)


logger = logging.getLogger("SLA")

logger.setLevel(logging.INFO)

formatter = logging.Formatter(
    "%(asctime)s | %(levelname)s | %(message)s"
)

arquivo_log = logging.FileHandler(
    "analise_sla.log",
    encoding="utf-8"
)

arquivo_log.setFormatter(formatter)

console = logging.StreamHandler()
console.setFormatter(formatter)

logger.addHandler(arquivo_log)
logger.addHandler(console)

def gerar_relatorio_encerrados(arquivo_saida, aba1, aba2, df_filtrado):
    # Gerar grupos em aba especifica
    relatorio = ExcelReport(
        arquivo=arquivo_saida,
        sheet_name=aba2,
        linhas_em_branco=2
        )

    # Retorna agrupada por 'IC afetado' e 'Atribuido a'
    df_agrupa1 = agrupar_ic_afetado(df_filtrado)
    df_agrupa2 = agrupar_resolvido_por(df_filtrado)

    relatorio.gerar([df_agrupa1, df_agrupa2])

    # Aplicar estilos na aba especifica
    aplicar_estilos_excel(arquivo_saida, aba1)

def gerar_relatorio_aberto(arquivo_saida, aba1, aba2, df_filtrado):
    # . Gerar grupos em aba especifica
    relatorio = ExcelReport(
        arquivo=arquivo_saida,
        sheet_name=aba2,
        linhas_em_branco=2
        )

    # . Retorna agrupada por 'IC afetado' e 'Atribuido a'
    df_agrupa1 = agrupar_ic_afetado(df_filtrado)
    df_agrupa2 = agrupar_atribuido_a(df_filtrado)

    # . Aplicar estilos na aba especifica
    aplicar_estilos_excel(arquivo_saida, aba1)

def gerar_relatorio_aberto_sctask(arquivo_saida, aba1, aba2, df_filtrado):
    #  Gerar grupos em aba especifica
    relatorio = ExcelReport(
        arquivo=arquivo_saida,
        sheet_name=aba2,
        linhas_em_branco=2
        )

    #  Retorna agrupada por 'IC afetado' e 'Atribuido a'
    df_agrupa1 = agrupar_por(df_filtrado,"Grupo de atribuição")
    df_agrupa2 = agrupar_por(df_filtrado,"Atribuição a")

    relatorio.gerar([df_agrupa1, df_agrupa2])
    # Aplicar estilos na aba especifica
    aplicar_estilos_excel(arquivo_saida, aba1)

# Função para aplicar cores com base nos valores da coluna "% SLA Resolution"
def colocar_cores_celulas(arquivo_saida, nome_aba, nome_coluna):
    wb = load_workbook(arquivo_saida)
    ws = wb[nome_aba]

    valores_sla = {
        "otimo": 0,
        "medio": 0,
        "atencao": 0,
        "fora": 0
    }

    cores = {
        "verde": PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE"),
        "amarelo": PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C"),
        "laranja": PatternFill("solid", start_color="FFD699", end_color="FFD699"),
        "vermelho": PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE"),
    }

    # Encontrar a coluna pelo nome
    col_idx = next(
        (
            cell.column
            for cell in ws[1]
            if cell.value == nome_coluna
        ),
        None
    )

    if col_idx is None:
        raise ValueError(
            f"Coluna '{nome_coluna}' não encontrada na aba '{nome_aba}'."
        )

    # Aplicar cores e contabilizar SLA
    for row in range(2, ws.max_row + 1):
        celula = ws.cell(row=row, column=col_idx)
        valor = celula.value

        if valor is None:
            continue

        if valor <= 45:
            celula.fill = cores["verde"]
            valores_sla["otimo"] += 1

        elif valor <= 50:
            celula.fill = cores["amarelo"]
            valores_sla["medio"] += 1

        elif valor <= 99:
            celula.fill = cores["laranja"]
            valores_sla["atencao"] += 1

        else:  # valor >= 100
            celula.fill = cores["vermelho"]
            valores_sla["fora"] += 1

    wb.save(arquivo_saida)

    print(f"Planilha '{arquivo_saida}' atualizada com sucesso.")

    df_sla = pd.DataFrame(
        valores_sla.items(),
        columns=["SLA", "Quantidade"]
    )

    print(df_sla)

    return valores_sla #, df_sla

# Comparar planilhas e gerar uma nova planilha com base nas configurações do arquivo JSON
def cruzar_planilhas(arquivo1, arquivo2, arquivo_saida, filtro, aba1, aba2):
    # 1. Ler os dados das planilhas para DataFrames
    df_plan1 = pd.read_excel(arquivo1)
    df_plan2 = pd.read_excel(arquivo2)
    df_vazio = pd.DataFrame()
    
    # 2. Renomear a coluna 0 para 'Chave' (caso o nome seja diferente)
    df_plan1.rename(columns={df_plan1.columns[0]: 'Numero'}, inplace=True)
    df_plan2.rename(columns={df_plan2.columns[0]: 'Numero'}, inplace=True)
    df_plan1.rename(columns={df_plan1.columns[2]: 'Aberto por'}, inplace=True)
    df_plan2.rename(columns={df_plan2.columns[6]: '% SLA Resolution'}, inplace=True)

    # 3. Selecionar apenas as colunas desejadas da planilha 2 Número e 6
    df_plan2 = df_plan2.iloc[:, [6,0]]  # Seleciona a coluna 0 (Chave) e a coluna 6 da planilha 2
    
    # 4. Filtra onde a coluna 5 contém o padrão Regex definido
    df_filtrado = df_plan1[df_plan1.iloc[:, 5].astype(str).str.contains(filtro, na=False, regex=True)]
   
    #df_filtrado.insert(0, '% SLA Response', [])
    # 5. Mesclar (Unir) os dados com base na coluna 0 (Chave)
    # Utilizando 'inner' para apenas onde há correspondência nas duas planilhas
    # Altere para 'outer', 'left' ou 'right' se desejar manter todos os dados independente de correspondência
    df_plan3 = pd.merge(df_plan2, df_filtrado, on='Numero', how='right')  # Mantendo todos os dados da planilha 1 e apenas os correspondentes da planilha 2

    # 5. Inserir a nova coluna no início (posição 0)
    df_plan3.insert(loc=0, column='% SLA Response', value='')

    # 6. Preencher os valores vazios da coluna 'NomeDaColuna' com 0
    df_plan3['% SLA Resolution'] = df_plan3['% SLA Resolution'].fillna(0)
    
    # 7. Gravar o arquivo Excel com duas abas
    with pd.ExcelWriter(arquivo_saida, engine='openpyxl') as writer:
           df_plan3.to_excel(writer, sheet_name=aba1, index=False)
           df_vazio.to_excel(writer, sheet_name=aba2, index=False) 

    # # 8. Gerar grupos em aba especifica
    # relatorio = ExcelReport(
    #     arquivo=arquivo_saida,
    #     sheet_name=aba2,
    #     linhas_em_branco=2
    #     )

    # # 10. Retorna agrupada por 'IC afetado' e 'Atribuido a'
    # df_agrupa1 = agrupar_ic_afetado(df_plan3)
    # df_agrupa2 = agrupar_atribuido_a(df_plan3)
    
    # relatorio.gerar([df_agrupa1, df_agrupa2])

    # 11. Aplicar estilos na aba especifica
    aplicar_estilos_excel(arquivo_saida, aba1)

    # 12. Exibir mensagem de sucesso 
    print(f"Arquivo '{arquivo_saida}' criado com sucesso!")    

    gerar_relatorio_aberto(arquivo_saida, aba1, aba2, df_plan3)

# processar planilhas com base nas configurações do arquivo JSON
def planilhas_encerrados(arquivo1, arquivo_saida, filtro, aba1, aba2): 
    # 1. Ler os dados da pplanilha 
    df_plan1 = pd.read_excel(arquivo1)
    df_vazio = pd.DataFrame()

    # 1.1 Renomear a coluna 2 para 'Aberto por'
    df_plan1.rename(columns={df_plan1.columns[2]: 'Aberto por'}, inplace=True)

    # 2. Filtrar os dados com base no valor da coluna 5
    df_filtrado = df_plan1[df_plan1.iloc[:, 5] == filtro]  # Filtra onde a coluna 5 é igual ao valor do filtro

    # 3. Gravar o arquivo Excel com duas abas
    with pd.ExcelWriter(arquivo_saida, engine='openpyxl') as writer:
           df_filtrado.to_excel(writer, sheet_name=aba1, index=False)
           df_vazio.to_excel(writer, sheet_name=aba2, index=False) 

    # 4. Aplicar estilos na aba especifica
    aplicar_estilos_excel(arquivo_saida, aba1)
    
    print(f"Arquivo '{data_atual}{arquivo_saida}' criado com sucesso!")

    gerar_relatorio_encerrados(arquivo_saida, aba1, aba2, df_filtrado)
    
# processar planilhas com base nas configurações do arquivo JSON
def planilhas_encerrados_aberto_sctask(arquivo1, arquivo_saida, aba1, aba2):
    # 1. Ler os dados da pplanilha 
    df_plan1 = pd.read_excel(arquivo1)
    df_vazio = pd.DataFrame()

    # 4. Gravar o arquivo Excel com duas abas
    with pd.ExcelWriter(arquivo_saida, engine='openpyxl') as writer:
           df_plan1.to_excel(writer, sheet_name=aba1, index=False)
           df_vazio.to_excel(writer, sheet_name=aba2, index=False) 

    gerar_relatorio_aberto_sctask(arquivo_saida, aba1, aba2, df_plan1)

    # 7. Aplicar estilos na aba especifica
    aplicar_estilos_excel(arquivo_saida, aba1)

    # 8. Exibir mensagem de sucesso
    print(f"Arquivo '{data_atual}{arquivo_saida}' criado com sucesso!") 
    
#  iniciar o processamento das planilhas com base nas configurações do arquivo JSON
def processar_configuracoes():
    # 1. ler o arquivo json
    with open(f'{pasta_trabalho}configuracao.json', 'r', encoding="utf-8") as f:
        configuracoes = json.load(f) 

    #2. criar uma nova planilha com base nas configurações do arquivo JSON
    for i in range(len(configuracoes)):

        # 2.1 Defina os nomes dos arquivos de entrada e saída com base nas configurações
        arquivo1 = f"{pasta_trabalho}{configuracoes[i]['ArquivoEntrada1']}"
        arquivo2 = f"{pasta_trabalho}{configuracoes[i]['ArquivoEntrada2']}"
        arquivo_saida = f"{pasta_trabalho}{data_atual}_{configuracoes[i]['arquivoSaida']}"
        filtro = configuracoes[i]['Filtro']
        JuntaArq = configuracoes[i]['JuntaArq']
        aba1 = configuracoes[i]['NovaAba1']
        aba2 = configuracoes[i]['NovaAba2']

        busca = arquivo1

        if "SCTASK" in busca:
            planilhas_encerrados_aberto_sctask(arquivo1, arquivo_saida, aba1, aba2)
            aplicar_cores_sla_totais(arquivo_saida, aba1, x)
        else:
            #2.2. Chame a função para cruzar as planilhas
            if JuntaArq == True:
                # 2.2.1. Chame a função para filtrar os dados da planilha 1 com base no filtro
                planilhas_encerrados(arquivo1, arquivo_saida, filtro, aba1, aba2) 
                # 2.2.2. aplicar_cores_condicionais(arquivo_saida,'% SLA Response',aba1)  
                #x=colocar_cores_celulas(arquivo_saida, aba1, '% SLA Resolution')
                # 2.2.3. aplicar_cores_condicionais(arquivo_saida,'% SLA Response',aba1)
                aplicar_cores_sla_totais(arquivo_saida, aba1,x)
            else:
                # 2.2.3. Chame a função para cruzar as planilhas
                cruzar_planilhas(arquivo1, arquivo2, arquivo_saida, filtro, aba1, aba2)   
                # 2.2.4. aplicar_cores_condicionais(arquivo_saida,'% SLA Response',aba1)  
                x= colocar_cores_celulas(arquivo_saida, aba1, '% SLA Resolution')
                # 2.2.5. aplicar_cores_condicionais(arquivo_saida,'% SLA Response',aba1)
                aplicar_cores_sla_totais(arquivo_saida, aba1,x)

# Criar planilha total com os painéis SLA 
def criar_planilha_total():
    # 1. Criar dois DataFrames vazios
    df_vazio = pd.DataFrame()
    arquivo_saida = "_Total.xlsx"

    # 2. Gravar o arquivo Excel com duas abas
    with pd.ExcelWriter(f'{pasta_trabalho}{data_atual}{arquivo_saida}', engine='openpyxl') as writer:
        df_vazio.to_excel(writer, sheet_name='Visão Geral', index=False)

    # 3. Aplicar estilos na aba especifica
    #aplicar_estilos_excel(f'{pasta_trabalho}{data_atual}{arquivo_saida}', 'Visão Geral')

    # 4. Carregar o arquivo Excel existente
    wb = load_workbook(f'{pasta_trabalho}{data_atual}{arquivo_saida}')
    ws = wb['Visão Geral']

    # 5. Criar os painéis SLA 
    aps.criar_painel_sla(
        ws,
        coluna_inicial=2,
        titulo="TRANSPORTE DOMÉSTICO",
        data_aberto=data_atual,
        data_encerrado=data_ontem,
        atribuicoes={},
        resolvidos={},
        sla_otimo=0,
        sla_medio=0,
        sla_atencao=0,
        sla_fora=0
    )

    # 6. Criar o painel SLA para Clearance - Transporte Internacional
    aps.criar_painel_sla(
        ws,
        coluna_inicial=12,
        titulo="CLEARANCE - TRANSPORTE INTERNACIONAL",
        data_aberto=data_atual,
        data_encerrado=data_ontem,
        atribuicoes={
            "Rafael Miranda Ferreira": 18,
            "Enoque Pinheiro": 10,
            "Victorio Bormolini": 1,
            "ANTONIO BARBOSA": 1
        },
        resolvidos={
            "ANTONIO BARBOSA": 1,
            "Victorio Bormolini": 1
        },
        sla_otimo= 0,#valores_sla["otimo"],
        sla_medio= 0,#valores_sla["medio"],
        sla_atencao=0,#valores_sla["atencao"],
        sla_fora=0,#valores_sla["fora"],
        detalhe_otimo={
            "LAC-BR-3-CSATH": 3,
            "LAC-BR-3-ClearanceFormal": 2
        },
        detalhe_medio={
            "LAC-BR-Clearance": 1,
            "LAC-BR-3-CSATH": 2
        },
        detalhe_atencao={
            "LAC-BR-3-CSATH": 5
        },
        detalhe_fora={
            "LAC-BR-Clearance": 16,
            "LAC-BR-3-CSATH": 1
        }
    )
    # 7. Salvar o arquivo final modificado
    wb.save(f'{pasta_trabalho}{data_atual}{arquivo_saida}') 

    # 8. Exibir mensagem de sucesso 
    print(f"Arquivo '{pasta_trabalho}{data_atual}{arquivo_saida}' criado com sucesso!") 
# agrupar a coluna 'IC afetado' e retornar a quantidade de ocorrências
def agrupar_ic_afetado(df):
    """
    Agrupa a coluna 'IC afetado' e retorna a quantidade de ocorrências.
    """
    resultado = (
        df.groupby("IC afetado")
        .size()
        .reset_index(name="Total")
    )

    linha_total = pd.DataFrame({
        "IC afetado": ["TOTAL GERAL"],
        "Total": [resultado["Total"].sum()]
    })

    resultado = pd.concat([resultado, linha_total], axis=0)
    
    return resultado
# agrupar a coluna passada como parametros nome_grupo e retornar a quantidade de ocorrências
def agrupar_por (df, nome_grupo):
    """
    Agrupa a coluna passada como parametros nome_grupo e retorna a quantidade de ocorrências.
    """
    resultado = (
        df.groupby(nome_grupo)
        .size()
        .reset_index(name="Total")
    )

    linha_total = pd.DataFrame({
        nome_grupo: ["TOTAL GERAL"],
        "Total": [resultado["Total"].sum()]
    })

    resultado = pd.concat([resultado, linha_total], axis=0)
    
    return resultado
# agrupar a coluna 'Atribuição a' e retornar a quantidade de ocorrências
def agrupar_atribuido_a(df):
    """
    Agrupa a coluna 'Atribuição a' e retorna a quantidade de ocorrências.
    """
    resultado = (
        df.groupby("Atribuição a")
        .size()
        .reset_index(name="Total")
    )

    linha_total = pd.DataFrame({
        "Atribuição a": ["TOTAL GERAL"],
        "Total": [resultado["Total"].sum()]
    })

    resultado = pd.concat([resultado, linha_total], axis=0)

    return resultado
# agrupar a coluna 'Resolvido por' e retornar a quantidade de ocorrências
def agrupar_resolvido_por(df):
    """
    Agrupa a coluna 'resolvido por' e retorna a quantidade de ocorrências.
    """
    resultado = (
        df.groupby("Resolvido por")
        .size()
        .reset_index(name="Total")
    )

    linha_total = pd.DataFrame({
        "Resolvido por": ["TOTAL GERAL"],
        "Total": [resultado["Total"].sum()]
    })

    resultado = pd.concat([resultado, linha_total], axis=0)

    return resultado
# aplicar estilos na planilha Excel
def aplicar_estilos_excel(arquivo_saida, aba):
    # 1. Carregar o arquivo Excel existente
    wb = load_workbook(arquivo_saida)
    ws = wb[aba]

    # Definir cores e estilos
    fill_titulo = PatternFill(
        fill_type="solid", start_color="1F4E79", end_color="1F4E79"
    )  # Azul escuro
    font_titulo = Font(color="FFFFFF", bold=True)

    fill_total = PatternFill(
        fill_type="solid", start_color="D9E1F2", end_color="D9E1F2"
    )  # Azul claro
    font_total = Font(bold=True)

    # Bordas com Side e Border
    borda_fina = Side(border_style="thin", color="000000")
    borda_dupla = Side(border_style="double", color="000000")

    borda_titulo = Border(
        top=borda_fina, bottom=borda_fina, left=borda_fina, right=borda_fina
    )

    # Aplicar estilos na linha do Título (Linha 1)
    for cell in ws[1]:
                cell.fill = fill_titulo
                cell.font = font_titulo
                cell.border = borda_titulo

    # Salvar o arquivo
    wb.save(arquivo_saida)
# aplicar valores SLA na planilha Excel
def aplicar_cores_sla_totais(arquivo_saida, aba, valores_sla=None):
    wb = load_workbook(arquivo_saida)

    ws = wb['SLA']
    wb.worksheets[1].title = 'SLA'  

    if aba in {"INT_aberto", "DOM_aberto", "CSATH_aberto"}:
        aplicar_sla(ws, valores_sla)
    else:
        print(f"Aba '{aba}' não reconhecida para aplicar os valores SLA.")

    wb.save(arquivo_saida) 

    print(f"Arquivo '{arquivo_saida}' com os valores SLA aplicados na aba '{aba}' com sucesso!")

def cabecalho_sla(ws):
    borda_fina = Side(
        border_style="thin", 
        color="000000"
        )
    # borda_dupla = Side(border_style="double", color="000000")
    borda = Border(
        left=borda_fina, 
        right=borda_fina, 
        top=borda_fina, 
        bottom=borda_fina
        )

    celula = ws["D1"]

    celula.value = "SLA"
    celula.fill = PatternFill(
        start_color="1F4E79",
        end_color="1F4E79",
        fill_type="solid"
    )
    celula.font = Font(color="FFFFFF", bold=True)
    celula.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    celula = ws["E1"]

    celula.value = "Quantidade"
    celula.fill = PatternFill(
        start_color="1F4E79",
        end_color="1F4E79",
        fill_type="solid"
    )
    celula.font = Font(color="FFFFFF", bold=True)
    celula.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    for row in ws.iter_rows(min_row=1, max_row=5, min_col=4, max_col=5):
        for cell in row:
            cell.border = borda

def aplicar_sla(ws, valores_sla):
    cabecalho_sla(ws)

    configuracoes = [
        ("D2", "Dentro do SLA", valores_sla["otimo"],   "C6EFCE"),
        ("D3", "Em andamento",  valores_sla["medio"],   "FFEB9C"),
        ("D4", "Em atenção",    valores_sla["atencao"], "FFD699"),
        ("D5", "Fora do SLA",   valores_sla["fora"],    "FFC7CE"),
    ]

    alinhamento = Alignment(horizontal="center", vertical="center")
    fonte = Font(bold=False)

    for celula_desc, texto, valor, cor in configuracoes:
        linha = celula_desc[1:]

        fill = PatternFill(
            start_color=cor,
            end_color=cor,
            fill_type="solid"
        )

        for celula, conteudo in [
            (f"D{linha}", texto),
            (f"E{linha}", valor),
        ]:
            ws[celula].value = conteudo
            ws[celula].fill = fill
            ws[celula].font = fonte
            ws[celula].alignment = alinhamento

# principal
def main():
    processar_configuracoes()     
    criar_planilha_total()
# executar o script principal
if __name__ == "__main__":
    main()



