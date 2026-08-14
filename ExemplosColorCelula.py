import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def colocar_core_celula(arquivo_saida,nome_aba,nome_coluna):
    # 2. Carregar o arquivo com openpyxl para manipular as cores
    wb = load_workbook(arquivo_saida)
    ws = wb.active
    ws.sheet_name=nome_aba

    # 3. Definir as cores de preenchimento (códigos Hexadecimais)
    verde = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')   # Verde claro
    amarelo = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid') # Amarelo claro
    laranja = PatternFill(start_color='FFD699', end_color='FFD699', fill_type='solid') # Laranja claro
    vermelho = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')# Vermelho claro

    # 4. Encontrar a coluna '% SLA Resolution' e aplicar as regras linha por linha
    col_idx = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == nome_coluna:
            col_idx = col
            break

    if col_idx:
        for row in range(2, ws.max_row + 1):
            celula = ws.cell(row=row, column=col_idx)
            valor = celula.value
            if valor is not None:
                if valor <= 45:
                    celula.fill = verde
                elif 46 <= valor <= 50:
                    celula.fill = amarelo
                elif 51 <= valor <= 99:
                    celula.fill = laranja
                elif valor >= 100:
                    celula.fill = vermelho

    # 5. Salvar o arquivo final modificado
    wb.save(arquivo_saida)
    print(F"Planilha {arquivo_saida} com as cores aplicadas!")

def main():
    colocar_core_celula(R'/Users/famj/projetos/analise_sla/20260730_CLEANCE_aberto.xlsx','CLEANCE_aberto','% SLA Resolution')

if __name__ == "__main__":
    main()

    