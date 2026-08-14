import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class ExcelReportHeaderOnly:

    def __init__(
        self,
        arquivo,
        sheet_name="Relatório",
        linhas_em_branco=2,
        cor_cabecalho="1F4E79"  # Azul escuro por padrão (Hex sem #)
    ):
        self.arquivo = arquivo
        self.sheet_name = sheet_name
        self.linhas_em_branco = linhas_em_branco

        # --- Estilo EXCLUSIVO do Cabeçalho ---
        self.fill_cabecalho = PatternFill(
            fill_type="solid",
            start_color=cor_cabecalho,
            end_color=cor_cabecalho
        )
        self.font_cabecalho = Font(color="FFFFFF", bold=True)
        self.alinhamento_cabecalho = Alignment(horizontal="center", vertical="center")

        # Bordas leves
        borda_fina = Side(border_style="thin", color="D3D3D3")
        self.borda_padrao = Border(
            top=borda_fina, bottom=borda_fina, left=borda_fina, right=borda_fina
        )

    def _garantir_arquivo_existe(self):
        """Cria um arquivo Excel básico se ele ainda não existir."""
        if not os.path.exists(self.arquivo):
            wb = Workbook()
            ws = wb.active
            ws.title = self.sheet_name
            wb.save(self.arquivo)

    def aplicar_estilos_tabela(self, ws, linha_cabecalho, n_linhas, n_colunas):
        """Aplica a cor e formatação apenas na linha do cabeçalho e bordas no corpo."""
        linha_fim = linha_cabecalho + n_linhas

        for row_idx, row in enumerate(
            ws.iter_rows(
                min_row=linha_cabecalho,
                max_row=linha_fim,
                min_col=1,
                max_col=n_colunas,
            ),
            start=linha_cabecalho,
        ):
            # 1. Se for a linha do CABEÇALHO (primeira linha do bloco)
            if row_idx == linha_cabecalho:
                for cell in row:
                    cell.fill = self.fill_cabecalho
                    cell.font = self.font_cabecalho
                    cell.alignment = self.alinhamento_cabecalho
                    cell.border = self.borda_padrao
            # 2. Demais linhas de dados (apenas borda leve)
            else:
                for cell in row:
                    cell.border = self.borda_padrao

    def gerar(self, dataframes):
        """Grava a lista de DataFrames e aplica o estilo do cabeçalho."""
        self._garantir_arquivo_existe()

        posicao_linha = 0
        blocos = []

        # 1. Escreve os DataFrames usando o Pandas
        with pd.ExcelWriter(
            self.arquivo, engine="openpyxl", mode="a", if_sheet_exists="overlay"
        ) as writer:

            for df in dataframes:
                df.to_excel(
                    writer,
                    sheet_name=self.sheet_name,
                    startrow=posicao_linha,
                    index=False,
                )

                blocos.append(
                    {
                        "linha_cabecalho": posicao_linha + 1,  # OpenPyXL usa índice 1
                        "n_linhas": len(df),
                        "n_colunas": len(df.columns),
                    }
                )

                # Avança a linha para o próximo bloco (dados + cabeçalho + espaço)
                posicao_linha += len(df) + 1 + self.linhas_em_branco

        # 2. Formata apenas os cabeçalhos usando OpenPyXL
        wb = load_workbook(self.arquivo)
        ws = wb[self.sheet_name]

        for bloco in blocos:
            self.aplicar_estilos_tabela(
                ws,
                bloco["linha_cabecalho"],
                bloco["n_linhas"],
                bloco["n_colunas"],
            )

        # 3. Ajuste automático da largura das colunas
        for col_idx, coluna in enumerate(ws.columns, start=1):
            largura_maxima = 0
            for cell in coluna:
                if cell.value is not None:
                    largura_maxima = max(largura_maxima, len(str(cell.value)))

            letra_coluna = get_column_letter(col_idx)
            ws.column_dimensions[letra_coluna].width = max(largura_maxima + 4, 12)

        wb.save(self.arquivo)